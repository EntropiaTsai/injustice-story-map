#!/usr/bin/env python3
"""
將臺灣轉型正義資料庫 Excel 轉成結構化 JSON。

輸出：data/processed/twtjdb_persons.json
每筆包含：
  - 基本個人資料
  - 終審判決摘要
  - location：從 f_txt 抽取「住○○」，僅保留台灣地名
  - location_status：「ready」（有地點）或「pending」（待補）
"""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("請先安裝 openpyxl：pip install openpyxl")
    sys.exit(1)

# ── 台灣縣市對照座標（WGS84 中心點） ──────────────────────────────────────
TAIWAN_LOCATIONS: dict[str, tuple[float, float]] = {
    # 直轄市
    "台北市": (25.0330, 121.5654),
    "臺北市": (25.0330, 121.5654),
    "新北市": (25.0120, 121.4651),
    "桃園市": (24.9937, 121.3010),
    "台中市": (24.1477, 120.6736),
    "臺中市": (24.1477, 120.6736),
    "台南市": (22.9999, 120.2270),
    "臺南市": (22.9999, 120.2270),
    "高雄市": (22.6273, 120.3014),
    # 縣
    "基隆市": (25.1276, 121.7392),
    "新竹市": (24.8066, 120.9686),
    "嘉義市": (23.4801, 120.4491),
    "宜蘭縣": (24.7021, 121.7378),
    "新竹縣": (24.8387, 121.0177),
    "苗栗縣": (24.5602, 120.8214),
    "彰化縣": (24.0684, 120.5820),
    "南投縣": (23.9609, 120.9718),
    "雲林縣": (23.7092, 120.4313),
    "嘉義縣": (23.4518, 120.2554),
    "屏東縣": (22.5519, 120.5487),
    "台東縣": (22.7972, 121.0713),
    "臺東縣": (22.7972, 121.0713),
    "花蓮縣": (23.9872, 121.6015),
    "澎湖縣": (23.5711, 119.5795),
    "金門縣": (24.4493, 118.3767),
    "連江縣": (26.1505, 119.9289),
    # 舊制縣名（已廢）
    "台北縣": (25.0120, 121.4651),
    "臺北縣": (25.0120, 121.4651),
    "高雄縣": (22.6273, 120.3014),
    "台南縣": (22.9999, 120.2270),
    "臺南縣": (22.9999, 120.2270),
    "桃園縣": (24.9937, 121.3010),
    "台中縣": (24.1477, 120.6736),
    "臺中縣": (24.1477, 120.6736),
    "台灣省": (23.6978, 120.9605),  # 省級 fallback
    "臺灣省": (23.6978, 120.9605),
    # 特殊
    "綠島": (22.6607, 121.4920),
    "火燒島": (22.6607, 121.4920),
    "龜山島": (24.8477, 121.9313),
}


def extract_residence(f_txt: str | None) -> str | None:
    """從判決書全文抽取「住○○」地名，回傳最早出現的縣市級地名。"""
    if not f_txt or f_txt == "暫無資料":
        return None

    # 優先比對「住○○縣/市」，再抓「籍設○○縣/市」
    # 字元類別排除冒號（：:），避免抓到「住址：福建省金門縣」之類格式
    patterns = [
        r"住([台臺]?灣?[省]?[^\s，。、（）():：\d]{2,6}(?:縣|市))",
        r"籍設([台臺]?灣?[省]?[^\s，。、（）():：\d]{2,6}(?:縣|市))",
    ]
    for pattern in patterns:
        m = re.search(pattern, f_txt)
        if m:
            # 額外清除殘留的前導冒號或空白（防禦性處理）
            return re.sub(r'^[：:\s]+', '', m.group(1)).strip()
    return None


def geocode(raw_location: str | None) -> tuple[float, float] | None:
    """將縣市名稱轉為座標；不在台灣則回傳 None。"""
    if not raw_location:
        return None
    for key, coords in TAIWAN_LOCATIONS.items():
        if key in raw_location:
            return coords
    return None


def clean_name(name: str | None) -> str:
    """保留主要姓名（頓號前），別名放到 aliases。
    清除規則：
    - 「空格＋中文數字」（如 " 一" " 二"）→ 案件序號，移除
    - 末尾阿拉伯數字（如 "陳光榮1"）→ 案件序號，移除
    - 中文數字在姓名中間或末尾無空格（如林守一、繆一良）→ 保留
    """
    if not name:
        return ""
    primary = str(name).split("、")[0].strip()
    primary = primary.split(" ")[0].strip()          # 移除「空格＋任何後綴」
    primary = re.sub(r'\d+$', '', primary).strip()   # 移除末尾阿拉伯數字
    return primary


def extract_aliases(name: str | None) -> list[str]:
    if not name:
        return []
    parts = str(name).replace(" ", "").split("、")
    return parts[1:] if len(parts) > 1 else []


def process(xlsx_path: str, output_path: str) -> None:
    print(f"讀取 {xlsx_path} ...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    def idx(col: str) -> int:
        return headers.index(col)

    # 取得需要的欄位索引
    fields = {
        "id": idx("id"),
        "name": idx("name"),
        "gender": idx("gender"),
        "birth_h": idx("birth_h"),
        "province": idx("province"),
        "city": idx("city"),
        "edu": idx("edu"),
        "occupation": idx("occupation"),
        "age": idx("age"),
        "f_authority": idx("f_authority"),
        "f_jyear": idx("f_jyear"),
        "f_penaltytxt": idx("f_penaltytxt"),
        "f_penalty": idx("f_penalty"),
        "f_death": idx("f_death"),
        "f_life": idx("f_life"),
        "f_term": idx("f_term"),
        "f_termy": idx("f_termy"),
        "f_group": idx("f_group"),
        "f_y": idx("f_y"),
        "f_m": idx("f_m"),
        "f_d": idx("f_d"),
        "f_txt": idx("f_txt"),
    }

    persons = []
    stats = {"total": 0, "ready": 0, "pending": 0, "mainland_only": 0}

    for row in ws.iter_rows(min_row=2, values_only=True):
        r = {k: row[v] for k, v in fields.items()}
        stats["total"] += 1

        raw_loc = extract_residence(r["f_txt"])
        coords = geocode(raw_loc)

        if raw_loc and coords is None:
            # 地名在中國大陸，不放台灣地圖
            stats["mainland_only"] += 1
            location_status = "mainland"
        elif coords:
            location_status = "ready"
            stats["ready"] += 1
        else:
            location_status = "pending"
            stats["pending"] += 1

        person: dict = {
            "id": str(r["id"]) if r["id"] else None,
            "name": clean_name(r["name"]),
            "aliases": extract_aliases(r["name"]),
            "gender": r["gender"],
            "birth_year_roc": r["birth_h"],
            "native_province": r["province"],
            "native_city": r["city"],
            "education": r["edu"],
            "occupation": r["occupation"],
            "age_at_arrest": r["age"],
            "final_judgment": {
                "authority": r["f_authority"],
                "year_roc": r["f_jyear"],
                "penalty_text": r["f_penaltytxt"],
                "has_death_penalty": r["f_death"] == "有",
                "has_life_sentence": r["f_life"] == "有",
                "has_term": r["f_term"] == "有",
                "term_years": r["f_termy"],
                "organization": r["f_group"],
                "date": {
                    "year": r["f_y"],
                    "month": r["f_m"],
                    "day": r["f_d"],
                },
            },
            "location_raw": raw_loc,
            "location_status": location_status,
        }

        if coords:
            person["lat"] = coords[0]
            person["lng"] = coords[1]

        persons.append(person)

    wb.close()

    # 輸出
    output = {
        "_meta": {
            "source": "臺灣轉型正義資料庫(14946筆)_20220420.xlsx",
            "total": stats["total"],
            "ready": stats["ready"],
            "pending": stats["pending"],
            "mainland_only": stats["mainland_only"],
        },
        "persons": persons,
    }

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n完成！")
    print(f"  總計：{stats['total']} 筆")
    print(f"  有地點（台灣）：{stats['ready']} 筆")
    print(f"  大陸地點：{stats['mainland_only']} 筆")
    print(f"  待補：{stats['pending']} 筆")
    print(f"  輸出：{output_path}")


if __name__ == "__main__":
    BASE = Path(__file__).parent.parent
    XLSX = BASE / "data/reference/twtjdb/臺灣轉型正義資料庫(14946筆)_20220420.xlsx"
    OUTPUT = BASE / "data/processed/twtjdb_persons.json"
    process(str(XLSX), str(OUTPUT))

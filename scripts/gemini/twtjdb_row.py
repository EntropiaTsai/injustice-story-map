"""
臺灣轉型正義資料庫 xlsx：讀取單列（供 material_from_twtjdb / twtjdb_structured 共用）。
"""
from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# 匯入 JSON notes_for_agents、純文字素材開頭等（與《編碼說明》一致，供 LLM 對齊語意）
TWTJDB_NOTES_FOR_AGENTS_SEMANTICS = (
    "【姓名欄】若原值僅為全形「－」或僅含連字／破折號等符號，表資料庫**未載明或占位**，"
    "勿當成本名的一部分。"
    "匯出已刪除字串末尾「空白＋一」之贅字（錄入殘留）。"
    "姓名字段末尾另可見「空白＋二」等，多與**案件序次**（第二案資料列）有關，勿併入別名判讀；全形「－」仍表未載明，勿與中文數字「一」混淆。"
    "【頓號】同一格內若有頓號「、」，**頓號後並非第二位受裁判人**；"
    "依《編碼說明》為該受裁判人之**別名**，或**檔案上所見之通用字**（含異體），仍屬同一人。"
    "【地理】`province`／`city` 在資料庫脈絡多為**籍貫**（祖籍／籍貫欄），**不是**事件發生地、"
    "拘捕地或審判地；主表若無單獨「案發／裁判地點」欄，請在缺漏清單標示**事件相關地點待補**，"
    "勿把籍貫直接當成地圖故事發生地。"
)

_NAME_PLACEHOLDER_ONLY = re.compile(
    r"^[\s\-－—─═~／/.。．・‧]+$"
)
_NAME_LITERAL_PLACEHOLDERS = frozenset(
    {"無", "同上", "同左", "N/A", "n/a", "NA", "unknown", "Unknown"},
)
# 取代後模型不會把符號誤當姓名；與《編碼說明》「－」表未載明一致
_NAME_PLACEHOLDER_REPLACEMENT = "（未載明；原資料庫為占位符，非姓名文字）"
# 姓名末尾「空白＋一」多為表單／排版殘留，非編碼說明所稱之別名用法
_NAME_TRAILING_SPACE_YI = re.compile(r"[\s　]+一\s*$")
# 末尾「空白＋中文數字」：全檔常標同一人之第 N 筆案件列（如「… 二」），非姓名本體
_CASE_ORDINAL_SUFFIX = re.compile(r"[\s　]+[一二三四五六七八九十百]+\s*$")


def _cleanup_twtjdb_name_trailing_noise(s: str) -> str:
    t = s.strip()
    t = _NAME_TRAILING_SPACE_YI.sub("", t)
    return t.strip()


def name_core_for_sibling_match(name: str) -> str:
    """比對「同人多案」時的姓名核心：剥除末尾案件序字與「空白＋一」贅字。"""
    t = name.strip()
    t = _NAME_TRAILING_SPACE_YI.sub("", t)
    t = _CASE_ORDINAL_SUFFIX.sub("", t)
    return t.strip()


def identity_key_for_merge(flat: dict[str, Any]) -> tuple[str, str, str, str] | None:
    """
    若可辨識，回傳 (姓名核心, 出生年字串, 籍貫省, 籍貫縣市)；否則 None（不併列）。
    須同時有姓名核心與籍貫省或縣市至少一項，避免僅姓名雷同誤併。
    """
    raw = flat.get("name")
    if not isinstance(raw, str):
        return None
    t = raw.strip()
    if not t or t == _NAME_PLACEHOLDER_REPLACEMENT:
        return None
    core = name_core_for_sibling_match(t)
    if not core:
        return None
    birth = ""
    bh = flat.get("birth_h")
    if bh is not None and str(bh).strip():
        birth = str(bh).strip()
    prov = str(flat.get("province", "")).strip() if flat.get("province") is not None else ""
    city = str(flat.get("city", "")).strip() if flat.get("city") is not None else ""
    if not prov and not city:
        return None
    return (core, birth, prov, city)


def _sibling_sort_key(item: tuple[int, dict[str, Any]]) -> tuple[int, int]:
    _, fl = item
    d1y = fl.get("d1_y")
    try:
        dy = int(d1y) if d1y is not None and str(d1y).strip().isdigit() else 9999
    except (TypeError, ValueError):
        dy = 9999
    rid = fl.get("id")
    try:
        iid = int(rid) if rid is not None and str(rid).strip().isdigit() else 0
    except (TypeError, ValueError):
        iid = 0
    return (dy, iid)


def collect_merge_siblings(
    xlsx: Path,
    anchor_excel_row: int,
    anchor_flat: dict[str, Any],
) -> list[tuple[int, dict[str, Any]]]:
    """
    掃描整份 xlsx，找出與 anchor 同一人之多筆案件列（姓名核心＋出生年＋籍貫皆一致）。
    若無法形成身分鍵，僅回傳錨點一筆。
    回傳依 d1_y、id 排序後之 (Excel 列號, flat) 列表。
    """
    key = identity_key_for_merge(anchor_flat)
    if key is None:
        return [(anchor_excel_row, dict(anchor_flat))]
    matched: list[tuple[int, dict[str, Any]]] = []
    for rnum, flat in iter_all_data_rows(xlsx):
        if identity_key_for_merge(flat) == key:
            matched.append((rnum, dict(flat)))
    matched.sort(key=_sibling_sort_key)
    return matched

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = (
    REPO_ROOT
    / "data"
    / "reference"
    / "twtjdb"
    / "臺灣轉型正義資料庫(14946筆)_20220420.xlsx"
)


def extract_by_excel_row(
    xlsx: Path,
    excel_row: int,
) -> tuple[tuple[Any, ...], tuple[Any, ...]]:
    """excel_row：Excel 列號，1=表頭，2=第一筆資料。"""
    if excel_row < 2:
        raise SystemExit("資料列請 >= 2（第 1 列為欄位名）。")
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(
            min_row=1,
            max_row=excel_row,
            values_only=True,
        )
        header = next(it)
        for _ in range(excel_row - 2):
            next(it)
        data = next(it)
        return header, data
    finally:
        wb.close()


def find_row_by_id(xlsx: Path, target_id: str) -> int:
    """回傳 Excel 列號；找不到則 exit 1。"""
    target = target_id.strip()
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(min_row=1, values_only=True)
        header = next(it)
        try:
            id_idx = list(header).index("id")
        except ValueError:
            id_idx = 0
        for rnum, row in enumerate(it, start=2):
            if row is None or id_idx >= len(row):
                continue
            cell = row[id_idx]
            if cell is None:
                continue
            if str(cell).strip() == target:
                return rnum
    finally:
        wb.close()
    print(f"找不到 id={target!r}。", file=sys.stderr)
    raise SystemExit(1)


def iter_all_data_rows(xlsx: Path):
    """
    單次開啟工作簿，逐列產出 (excel_row, flat)。
    第 1 列為表頭；跳過 flat 為空的列。
    """
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(min_row=1, values_only=True)
        header_row = next(it, None)
        if header_row is None:
            return
        header = header_row
        for rnum, row in enumerate(it, start=2):
            if row is None:
                continue
            flat = row_to_dict(header, row)
            if not flat:
                continue
            yield rnum, flat
    finally:
        wb.close()


def row_to_dict(headers: tuple[Any, ...], values: tuple[Any, ...]) -> dict[str, Any]:
    """略過空值；鍵為欄位名字串。"""
    out: dict[str, Any] = {}
    for h, v in zip(headers, values):
        if h is None:
            continue
        key = str(h).strip()
        if not key:
            continue
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        if isinstance(v, str):
            out[key] = v.strip()
        else:
            out[key] = v
    return normalize_twtjdb_flat(out)


def _is_twtjdb_name_placeholder(value: str) -> bool:
    t = value.strip()
    if not t:
        return False
    if t in _NAME_LITERAL_PLACEHOLDERS:
        return True
    return bool(_NAME_PLACEHOLDER_ONLY.fullmatch(t))


def normalize_twtjdb_flat(flat: dict[str, Any]) -> dict[str, Any]:
    """
    將資料庫慣例轉成對 LLM 友善的表示：
    - 先去除姓名末尾「空白＋一」贅字（錄入殘留，非別名規則）。
    - 僅占位的 name 改為明確「未載明」句。
    含頓號「、」的別名／通用字區段不另刪除（語意見匯出說明）。
    """
    out = dict(flat)
    if "name" in out:
        v = out["name"]
        if isinstance(v, str):
            v = _cleanup_twtjdb_name_trailing_noise(v)
            if _is_twtjdb_name_placeholder(v):
                out["name"] = _NAME_PLACEHOLDER_REPLACEMENT
            else:
                out["name"] = v
    return out


def twtjdb_semantics_md_lines() -> list[str]:
    """插入 structured 匯出 Markdown（§2 區塊內）。"""
    return [
        "### 資料庫慣例（`name` 欄，務必遵守）",
        "",
        "- 《臺灣轉型正義資料庫編碼說明》：欄位若**僅**為全形「**－**」或僅含連字、破折號等符號，表**未載明**或無從查考，**不要**把符號當成本名的一部分。（匯出工具已將「僅符號」改寫為「未載明」句，若仍見符號請依編碼說明判讀。）",
        "- 同一格內若有**頓號「、」**：頓號**後面不是第二位受裁判人**；說明載明，頓號後為該受裁判人之**別名**，或於**檔案上出現的通用字**（含異體寫法）。敘事與結構化時應視為**同一人**之稱呼變體，勿拆成兩人。",
        "- 匯出工具會自動刪除 `name` **字串最末尾**的「空白＋一」（如 `王建 一`→`王建`），此為常見錄入贅字，**不**適用於頓號後別名規則。",
        "",
        "### 地理欄位（`province`／`city` 勿當案發地）",
        "",
        "- 資料庫常見欄位 `province`、`city` 依編碼脈絡為**籍貫（省／縣市）**，用於出身背景，**不代表**本案事件發生地、逮捕處所或法院所在地。",
        "- 若素材未另載明與本案相關之地點，正規化時應**分欄**標示「籍貫」與「事件／裁判相關地點（待補）」，勿混用；地圖展示用座標應以**事件相關地點**為優先，缺件時交 §5 地理專員提出查證路徑與保守描述。",
        "",
    ]


def twtjdb_semantics_plaintext_block() -> str:
    """插入 material_from_twtjdb 純文字開頭。"""
    return (
        "【姓名欄慣例（依《臺灣轉型正義資料庫編碼說明》）】\n"
        "- 「name」僅為全形「－」或僅符號者：表未載明；本匯出已將「僅符號」改寫為明確句，勿當成本名。\n"
        "- 同一格內頓號「、」後：為**同一人**之別名或檔案上通用字，**非**第二位受裁判人。\n"
        "- 末尾「空白＋一」已由匯出刪除（錄入贅字）。\n"
        "\n"
        "【地理】`province`／`city` 多為籍貫，**不是**案發地或法院地；若無另載地點請標「事件相關地點待補」。\n"
        "\n"
    )


def group_by_field_prefix(flat: dict[str, Any]) -> dict[str, dict[str, Any]]:
    """
    依欄位名稱第一層前綴分組（例：d1_crime1 → 群組 d1；無底線 → _base）。
    """
    groups: dict[str, dict[str, Any]] = {}
    for k, v in flat.items():
        if "_" in k:
            g = k.split("_", 1)[0]
        else:
            g = "_base"
        groups.setdefault(g, {})[k] = v
    return dict(sorted(groups.items(), key=lambda x: x[0]))

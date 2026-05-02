"""
讀取《臺灣轉型正義資料庫編碼說明》xlsx，建欄位代碼 → 中文標題與說明，供匯出腳本併入素材。
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from twtjdb_row import REPO_ROOT

DEFAULT_ENCODING_XLSX = (
    REPO_ROOT
    / "data"
    / "reference"
    / "twtjdb"
    / "臺灣轉型正義資料庫編碼說明_20210226.xlsx"
)

_FIELD_KEY_RE = re.compile(r"^[a-z][a-z0-9_]*$")
# 解析邏輯變更時遞增，避免程式重載前誤用舊快取
_PARSER_REVISION = 2

# 主表／純文字摘要欄避免單格過長
_MAX_NOTES_INLINE = 420


def parse_encoding_xlsx(path: Path) -> dict[str, dict[str, str]]:
    """
    自編碼說明各工作表解析欄位代碼。
    回傳 field_key -> {label_zh, description, sheet}
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, dict[str, str]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        pending_label: str | None = None
        pending_desc: list[str] = []

        for row in ws.iter_rows(values_only=True):
            cells = list(row)
            a = cells[0] if cells else None
            b = cells[1] if len(cells) > 1 else None

            if (
                isinstance(a, (int, float))
                and float(a).is_integer()
                and b is not None
                and str(b).strip()
            ):
                pending_label = str(b).strip()
                pending_desc = []
                continue

            if pending_label is None:
                continue

            field_tokens: list[str] = []
            same_row_text: list[str] = []
            for cell in cells:
                if cell is None:
                    continue
                s = str(cell).strip()
                if _FIELD_KEY_RE.match(s):
                    field_tokens.append(s)
                elif s != pending_label:
                    same_row_text.append(s)

            if field_tokens:
                desc_text = "\n".join(
                    x.strip() for x in pending_desc if x and str(x).strip()
                ).strip()
                row_extra = "\n".join(same_row_text).strip()
                if row_extra:
                    desc_text = (desc_text + "\n" + row_extra).strip() if desc_text else row_extra
                for fk in field_tokens:
                    entry = {
                        "label_zh": pending_label,
                        "description": desc_text,
                        "sheet": sheet_name,
                    }
                    if fk in result:
                        prev = result[fk]
                        nd = desc_text
                        if nd and nd not in prev["description"]:
                            entry = {
                                "label_zh": prev["label_zh"],
                                "description": (
                                    prev["description"]
                                    + "\n\n── "
                                    + sheet_name
                                    + " ──\n"
                                    + nd
                                ).strip(),
                                "sheet": prev["sheet"] + "；" + sheet_name,
                            }
                        else:
                            entry = {
                                "label_zh": prev["label_zh"],
                                "description": prev["description"],
                                "sheet": prev["sheet"] + "；" + sheet_name,
                            }
                    result[fk] = entry
                pending_desc = []
            else:
                parts: list[str] = []
                if b is not None and str(b).strip():
                    bs = str(b).strip()
                    if not _FIELD_KEY_RE.match(bs):
                        parts.append(bs)
                if len(cells) > 2 and cells[2] is not None and str(cells[2]).strip():
                    cs = str(cells[2]).strip()
                    if not _FIELD_KEY_RE.match(cs):
                        parts.append(cs)
                if parts:
                    pending_desc.append(" ".join(parts))

    wb.close()
    return result


_catalog_mtime: float | None = None
_catalog_data: dict[str, dict[str, str]] | None = None
_catalog_rev: int | None = None


def load_field_encoding_catalog(path: Path | None = None) -> dict[str, dict[str, str]]:
    """依檔案 mtime 與解析版本快取；檔案不存在回傳空 dict。"""
    global _catalog_mtime, _catalog_data, _catalog_rev
    p = (path or DEFAULT_ENCODING_XLSX).resolve()
    if not p.is_file():
        return {}
    mtime = p.stat().st_mtime
    if (
        _catalog_data is not None
        and _catalog_mtime == mtime
        and _catalog_rev == _PARSER_REVISION
    ):
        return _catalog_data
    _catalog_data = parse_encoding_xlsx(p)
    _catalog_mtime = mtime
    _catalog_rev = _PARSER_REVISION
    return _catalog_data


def subset_for_flat(
    flat: dict[str, Any], catalog: dict[str, dict[str, str]]
) -> dict[str, dict[str, str]]:
    """僅保留本列出現的欄位。"""
    return {k: dict(v) for k, v in catalog.items() if k in flat}


def truncate_for_table_note(description: str, max_len: int = _MAX_NOTES_INLINE) -> str:
    t = description.strip()
    if len(t) <= max_len:
        return t
    return t[: max_len - 1] + "…"


def render_flat_encoding_text(
    flat: dict[str, Any],
    catalog: dict[str, dict[str, str]],
    *,
    max_desc: int = 600,
) -> str:
    """純文字附錄：本列欄位在編碼說明中的標題與說明。"""
    lines = [
        "【欄位說明（臺灣轉型正義資料庫編碼說明 xlsx，僅列本列出現之欄位代碼）】",
        "",
    ]
    for key in sorted(flat.keys()):
        c = catalog.get(key)
        if not c:
            lines.append(f"- `{key}`：（編碼說明檔未收錄此代碼，請以欄位語意謹慎處理）")
            lines.append("")
            continue
        lz = c.get("label_zh", "").strip()
        sh = c.get("sheet", "").strip()
        desc = c.get("description", "").strip()
        lines.append(f"- `{key}` → {lz}（表：{sh}）")
        if desc:
            d = desc if len(desc) <= max_desc else desc[: max_desc - 1] + "…"
            for para in d.split("\n"):
                if para.strip():
                    lines.append(f"  {para.strip()}")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"
